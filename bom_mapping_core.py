import io
from typing import List, Tuple
import pandas as pd

def normalize_text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    if text.lower() == 'nan':
        return ''
    return text

def normalize_key(value):
    return normalize_text(value).upper()

def read_excel_safely(uploaded_file, sheet_name):
    filename = uploaded_file.name.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
        engine = 'openpyxl'
    elif filename.endswith('.xls'):
        engine = 'xlrd'
    else:
        raise ValueError(
            f'Unsupported file type: {uploaded_file.name}. Please upload .xlsx, .xlsm, or .xls'
        )

    uploaded_file.seek(0)
    try:
        return pd.read_excel(uploaded_file, sheet_name=sheet_name, engine=engine)
    except Exception as e:
        raise ValueError(
            f'Cannot find sheet "{sheet_name}" in file "{uploaded_file.name}". '
            'Please rename or provide the correct mapping sheet.'
        ) from e

CATEGORY_RULES = [
    ('023', ['resistor network', 'res network', 'ressip', 'array resistor', 'res net']),
    ('024', ['rc network']),
    ('022', ['resistor', 'res,', 'res ']),
    ('021', ['capacitor', 'cap,', 'cap ']),
    ('025', ['led', 'lcd', 'optical']),
    ('026', ['xcvr', 'rcvr', 'transceiver', 'receiver module']),
    ('027', ['analog ic', 'op amp', 'adc', 'dac', 'comparator', 'regulator']),
    ('029', ['digital ic', 'logic', 'buffer', 'flip-flop', 'mux', 'demux', 'gate']),
    ('030', ['eprom', 'eeprom', 'flash', 'fpga', 'mcu', 'microcontroller', 'cpu', 'special ic', 'hybrid', 'gate array']),
    ('031', ['ptc']),
    ('032', ['pal', 'prom', 'programmed device', 'memory']),
    ('033', ['delayline', 'delay line']),
    ('034', ['crystal', 'oscillator']),
    ('035', ['diode', 'transistor', 'rectifier', 'varistor', 'mosfet', 'xstr', 'zener']),
    ('036', ['thermistor']),
    ('037', ['inductor', 'ind,', 'ind ', 'ferrite', 'choke', 'tvs', 'filter', 'transformer', 'xfmr', 'bead']),
    ('038', ['fuse', 'circuit breaker', 'holder']),
    ('039', ['relay']),
    ('041', ['switch']),
    ('043', ['header', 'receptacle', 'shunt', 'pin']),
    ('044', ['connector', 'conn', 'socket', 'adapter', 'terminal', 'term']),
    ('045', ['screw', 'nut', 'washer', 'fastener', 'insul', 'insulator', 'standoff']),
    ('046', ['cable', 'wire', 'cord']),
    ('047', ['hardware', 'chassis', 'bracket', 'guard', 'cover', 'mounting', 'panel', 'plate', 'guide', 'lock', 'emi']),
    ('048', ['power supply', 'pwrsply', 'fan', 'battery']),
    ('049', ['heatsink', 'heat sink']),
]

CATEGORY_DESCRIPTION = {
    '015': 'PCB FAB',
    '021': 'CAPACITOR',
    '022': 'RESISTOR',
    '023': 'RESISTOR NETWORK, RESSIP',
    '024': 'RC NETWORK',
    '025': 'OPTICAL COMPONENT, LED, LCD',
    '026': 'ELECTRONIC MODULE, XCVR, RCVR',
    '027': 'IC, ANALOG',
    '029': 'IC, DIGITAL STANDARD',
    '030': 'IC, OTHER / SPECIAL, HYBRID, EPROM, EPROM SET, GATE ARRAY',
    '031': 'PTC',
    '032': 'MEM, PROG DEVICE, PAL, PROM',
    '033': 'DELAYLINE',
    '034': 'CRYSTAL, OSCILLATOR',
    '035': 'DISCRETE SEMI, DIODE, XSTR, RCTFR, VARISTOR',
    '036': 'THERMISTORS',
    '037': 'FILTER, INDUCTOR, FERRITE, XFMR, TVS, CHOKE',
    '038': 'FUSE, CIRCUIT BREAKER, HOLDER',
    '039': 'RELAY',
    '041': 'SWITCH',
    '043': 'HEADER, RECEPTACLE, SHUNT, PIN',
    '044': 'CONNECTOR, SOCKET, ADAPTER, TERM',
    '045': 'SCREW, NUT, FASTENER, WASHER, INSUL',
    '046': 'CABLE, WIRE, CORD',
    '047': 'HARDWARE, CHASSIS, BRACKET, GUARD, COVER, MOUNTING EARS, PANEL, PLATE, GUIDE, LOCK, EMI',
    '048': 'PWRSPLY, FAN, BAT',
    '049': 'HEATSINK MODULE',
    '051': 'ASSY PRODUCTION SUPPLY, COMPOUND, LOCTITE GLUE, KAPTON TAPE',
}

def suggest_part_number_category(description, location, bom_mpn_list):
    desc = normalize_text(description).lower()
    loc = normalize_text(location).upper()
    mpn_text = normalize_text(bom_mpn_list).lower()
    text = ' '.join([desc, normalize_text(location).lower(), mpn_text])

    if (
        'underfill' in text or 'glue' in text or 'adhesive' in text or 'epoxy' in text
        or 'loctite' in text or 'kapton' in text or 'tape' in text or 'compound' in text
    ):
        return '051', CATEGORY_DESCRIPTION['051'], 'Description/MPN suggests glue/underfill/compound'

    if (
        desc == 'pcb' or desc.startswith('pcb') or ' pcb' in f' {desc}'
        or loc.startswith('PCB') or 'bare board' in text or 'printed circuit board' in text
    ):
        return '015', CATEGORY_DESCRIPTION['015'], 'Description/location suggests PCB'

    if desc.startswith('res') or ' res,' in f' {desc}' or 'resistor' in desc:
        return '022', CATEGORY_DESCRIPTION['022'], 'Description suggests resistor'
    if desc.startswith('cap') or ' cap,' in f' {desc}' or 'capacitor' in desc:
        return '021', CATEGORY_DESCRIPTION['021'], 'Description suggests capacitor'
    if desc.startswith('ind') or ' ind,' in f' {desc}' or 'inductor' in desc:
        return '037', CATEGORY_DESCRIPTION['037'], 'Description suggests inductor/filter'
    if desc.startswith('conn') or ' conn' in f' {desc}' or 'connector' in desc:
        return '044', CATEGORY_DESCRIPTION['044'], 'Description suggests connector'
    if 'resistor network' in desc or 'ressip' in desc or desc.startswith('rn'):
        return '023', CATEGORY_DESCRIPTION['023'], 'Description suggests resistor network'
    if 'rc network' in desc:
        return '024', CATEGORY_DESCRIPTION['024'], 'Description suggests RC network'

    if loc.startswith('RN') or loc.startswith('RA'):
        return '023', CATEGORY_DESCRIPTION['023'], 'Location prefix suggests resistor network'
    if loc.startswith('R'):
        return '022', CATEGORY_DESCRIPTION['022'], 'Location prefix suggests resistor'
    if loc.startswith('C'):
        return '021', CATEGORY_DESCRIPTION['021'], 'Location prefix suggests capacitor'
    if loc.startswith('L') or loc.startswith('FB') or loc.startswith('T'):
        return '037', CATEGORY_DESCRIPTION['037'], 'Location prefix suggests inductor/filter'
    if loc.startswith('D') or loc.startswith('Q') or loc.startswith('CR'):
        return '035', CATEGORY_DESCRIPTION['035'], 'Location prefix suggests discrete semiconductor'
    if loc.startswith('F'):
        return '038', CATEGORY_DESCRIPTION['038'], 'Location prefix suggests fuse'
    if loc.startswith('K'):
        return '039', CATEGORY_DESCRIPTION['039'], 'Location prefix suggests relay'
    if loc.startswith('J') or loc.startswith('P') or loc.startswith('CN'):
        return '044', CATEGORY_DESCRIPTION['044'], 'Location prefix suggests connector'
    if loc.startswith('H'):
        return '043', CATEGORY_DESCRIPTION['043'], 'Location prefix suggests header/pin'
    if loc.startswith('U') or loc.startswith('IC'):
        return '030', CATEGORY_DESCRIPTION['030'], 'Location prefix suggests IC'
    if loc.startswith('Y') or loc.startswith('X'):
        return '034', CATEGORY_DESCRIPTION['034'], 'Location prefix suggests crystal/oscillator'
    if loc.startswith('SW') or loc.startswith('S'):
        return '041', CATEGORY_DESCRIPTION['041'], 'Location prefix suggests switch'

    for code, keywords in CATEGORY_RULES:
        for kw in keywords:
            if kw in text:
                return code, CATEGORY_DESCRIPTION.get(code, ''), f'Keyword match: {kw}'

    return '', 'Needs Review', 'Unable to classify from description/location/MPN'

def extract_bom_mpn_pairs(row_values):
    pairs = []
    primary_mfg = normalize_text(row_values[4]) if len(row_values) > 4 else ''
    primary_mpn = normalize_text(row_values[5]) if len(row_values) > 5 else ''
    if primary_mfg or primary_mpn:
        pairs.append((primary_mfg, primary_mpn, 'Primary'))

    alt_values = row_values[6:] if len(row_values) > 6 else []
    alt_index = 1
    for i in range(0, len(alt_values), 2):
        alt_mfg = normalize_text(alt_values[i]) if i < len(alt_values) else ''
        alt_mpn = normalize_text(alt_values[i + 1]) if i + 1 < len(alt_values) else ''
        if alt_mfg or alt_mpn:
            pairs.append((alt_mfg, alt_mpn, f'Alt {alt_index}'))
            alt_index += 1
    return pairs

def read_original_bom(uploaded_file, sheet_name='BOM', data_start_row=2):
    uploaded_file.seek(0)
    filename = uploaded_file.name.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
        engine = 'openpyxl'
    elif filename.endswith('.xls'):
        engine = 'xlrd'
    else:
        raise ValueError(
            f'Unsupported file type: {uploaded_file.name}. Please upload .xlsx, .xlsm, or .xls'
        )

    skiprows = max(data_start_row - 1, 0)
    try:
        df = pd.read_excel(
            uploaded_file,
            sheet_name=sheet_name,
            header=None,
            skiprows=skiprows,
            engine=engine,
        )
    except Exception as e:
        raise ValueError(
            f'Cannot find sheet "{sheet_name}" in file "{uploaded_file.name}".'
        ) from e

    if df.shape[1] < 6:
        raise ValueError('Original BOM must include at least columns A through F.')

    base_rows = []
    mpn_rows = []
    for _, row in df.iterrows():
        row_values = row.tolist()
        cpn = normalize_text(row_values[0]) if len(row_values) > 0 else ''
        description = normalize_text(row_values[1]) if len(row_values) > 1 else ''
        qty = normalize_text(row_values[2]) if len(row_values) > 2 else ''
        location = normalize_text(row_values[3]) if len(row_values) > 3 else ''

        if not any([cpn, description, qty, location]):
            continue

        base_rows.append({
            'Customer_CPN': cpn,
            'Description': description,
            'Qty_Per_Board': qty,
            'Location': location,
        })

        for mfg, mpn, source in extract_bom_mpn_pairs(row_values):
            if mfg or mpn:
                mpn_rows.append({
                    'Customer_CPN': cpn,
                    'Description': description,
                    'Qty_Per_Board': qty,
                    'Location': location,
                    'BOM_MFG': mfg,
                    'BOM_MPN': mpn,
                    'Source': source,
                })

    base_df = pd.DataFrame(base_rows).drop_duplicates().reset_index(drop=True)
    mpn_df = pd.DataFrame(mpn_rows).drop_duplicates().reset_index(drop=True)
    if base_df.empty:
        raise ValueError('No usable Original BOM data found.')

    return base_df, mpn_df

def read_organized_cpn_mapping(uploaded_file, sheet_name='SPN-CPN Mapping'):
    df = read_excel_safely(uploaded_file, sheet_name=sheet_name)
    first_row = [normalize_key(v) for v in df.iloc[0].tolist()] if len(df) > 0 else []
    if len(first_row) >= 2 and first_row[0] == 'SPN' and first_row[1] == 'CPN':
        df = df.iloc[1:].reset_index(drop=True)

    if df.shape[1] < 2:
        raise ValueError('Sheet "SPN-CPN Mapping" must have at least 2 columns (A=SPN, B=CPN).')

    out = df.iloc[:, [0, 1]].copy()
    out.columns = ['SPN', 'CPN']
    out['SPN'] = out['SPN'].apply(normalize_text)
    out['CPN'] = out['CPN'].apply(normalize_text)
    out = out[(out['SPN'] != '') & (out['CPN'] != '')].drop_duplicates().reset_index(drop=True)
    return out

def read_organized_mpn_mapping(uploaded_file, sheet_name='SPN-MPN Mapping'):
    df = read_excel_safely(uploaded_file, sheet_name=sheet_name)
    first_row = [normalize_key(v) for v in df.iloc[0].tolist()] if len(df) > 0 else []
    if len(first_row) >= 3 and first_row[0] == 'SPN' and first_row[2] == 'MPN':
        df = df.iloc[1:].reset_index(drop=True)

    if df.shape[1] < 3:
        raise ValueError('Sheet "SPN-MPN Mapping" must have at least 3 columns (A=SPN, B=MFG, C=MPN).')

    out = df.iloc[:, [0, 1, 2]].copy()
    out.columns = ['SPN', 'System_MFG', 'System_MPN']
    out['SPN'] = out['SPN'].apply(normalize_text)
    out['System_MFG'] = out['System_MFG'].apply(normalize_text)
    out['System_MPN'] = out['System_MPN'].apply(normalize_text)
    out = out[(out['SPN'] != '') & (out['System_MPN'] != '')].drop_duplicates().reset_index(drop=True)
    return out

def map_cpn_to_spn(original_base_df, original_mpn_df, cpn_mapping_df, system_mpn_df):
    bom_groups = {}
    for _, row in original_mpn_df.iterrows():
        cpn = normalize_text(row['Customer_CPN'])
        mpn = normalize_key(row['BOM_MPN'])
        if cpn not in bom_groups:
            bom_groups[cpn] = set()
        if mpn:
            bom_groups[cpn].add(mpn)

    sys_groups = {}
    for _, row in system_mpn_df.iterrows():
        spn = normalize_text(row['SPN'])
        mpn = normalize_key(row['System_MPN'])
        # 被 cross 的 MPN 視為系統已無此料，不納入 SPN 挑選的 overlap 計算。
        if row.get('Is_Crossed', False):
            continue
        if spn not in sys_groups:
            sys_groups[spn] = set()
        if mpn:
            sys_groups[spn].add(mpn)

    mapping = cpn_mapping_df.copy()
    mapping['CPN_KEY'] = mapping['CPN'].apply(normalize_key)

    cpn_to_spns = (
        mapping.groupby('CPN_KEY')['SPN']
        .agg(lambda x: sorted(set(normalize_text(v) for v in x if normalize_text(v))))
        .to_dict()
    )

    result_rows = []
    for _, row in original_base_df.iterrows():
        customer_cpn = normalize_text(row['Customer_CPN'])
        cpn_key = normalize_key(customer_cpn)
        description = row.get('Description', '')
        qty = row.get('Qty_Per_Board', '')
        location = row.get('Location', '')
        candidate_spns = cpn_to_spns.get(cpn_key, [])
        bom_set = bom_groups.get(customer_cpn, set())

        if not candidate_spns:
            result_rows.append({
                'Customer_CPN': customer_cpn,
                'Description': description,
                'Qty_Per_Board': qty,
                'Location': location,
                'SPN': '',
                'Candidate_SPNs': '',
                'Selection_Status': 'Missing SPN',
                'Best_Overlap_Count': 0,
            })
            continue

        if len(candidate_spns) == 1:
            selected_spn = candidate_spns[0]
            overlap_count = len(bom_set & sys_groups.get(selected_spn, set()))
            result_rows.append({
                'Customer_CPN': customer_cpn,
                'Description': description,
                'Qty_Per_Board': qty,
                'Location': location,
                'SPN': selected_spn,
                'Candidate_SPNs': selected_spn,
                'Selection_Status': 'Unique match',
                'Best_Overlap_Count': overlap_count,
            })
            continue

        scored = []
        for spn in candidate_spns:
            system_set = sys_groups.get(spn, set())
            overlap = len(bom_set & system_set)
            scored.append((spn, overlap))

        scored = sorted(scored, key=lambda x: x[1], reverse=True)
        best_spn, best_score = scored[0]
        top_ties = [spn for spn, score in scored if score == best_score]

        if len(top_ties) == 1:
            selection_status = 'Auto-selected by MPN overlap'
            selected_spn = best_spn
        else:
            selection_status = 'Ambiguous - same overlap'
            selected_spn = ''

        result_rows.append({
            'Customer_CPN': customer_cpn,
            'Description': description,
            'Qty_Per_Board': qty,
            'Location': location,
            'SPN': selected_spn,
            'Candidate_SPNs': ' / '.join(candidate_spns),
            'Selection_Status': selection_status,
            'Best_Overlap_Count': best_score,
        })

    return pd.DataFrame(result_rows)

def build_spn_detail(candidate_spns_text, bom_set, sys_groups):
    candidate_spns = [s.strip() for s in normalize_text(candidate_spns_text).split(' / ') if s.strip()]
    blocks = []
    for spn in candidate_spns:
        system_set = sys_groups.get(spn, set())
        missing = sorted(bom_set - system_set)
        extra = sorted(system_set - bom_set)

        block = f'[{spn}]\n'
        block += f"System: {' / '.join(sorted(system_set)) if system_set else '-'}\n"
        block += f"Missing: {' / '.join(missing) if missing else '-'}\n"
        block += f"Extra: {' / '.join(extra) if extra else '-'}"
        blocks.append(block)

    return '\n\n'.join(blocks)

def build_mpn_compare(mapped_df, original_mpn_df, system_mpn_df):
    bom_groups = {}
    for _, row in original_mpn_df.iterrows():
        cpn = normalize_text(row["Customer_CPN"])
        mpn = normalize_key(row["BOM_MPN"])
        if cpn not in bom_groups:
            bom_groups[cpn] = set()
        if mpn:
            bom_groups[cpn].add(mpn)

    sys_groups = {}
    crossed_groups = {}
    for _, row in system_mpn_df.iterrows():
        spn = normalize_text(row["SPN"])
        mpn = normalize_key(row["System_MPN"])
        if not mpn:
            continue
        # 被 cross 的 MPN 另外存到 crossed_groups，不算進系統有效 MPN。
        if row.get("Is_Crossed", False):
            crossed_groups.setdefault(spn, set()).add(mpn)
        else:
            sys_groups.setdefault(spn, set()).add(mpn)

    compare_rows = []
    for _, row in mapped_df.iterrows():
        cpn = normalize_text(row["Customer_CPN"])
        spn = normalize_text(row.get("SPN", ""))
        desc = normalize_text(row.get("Description", ""))
        loc = normalize_text(row.get("Location", ""))
        selection_status = normalize_text(row.get("Selection_Status", ""))
        candidate_spns = normalize_text(row.get("Candidate_SPNs", ""))
        bom_set = bom_groups.get(cpn, set())

        # 只有 Ambiguous 才顯示 detail
        if selection_status == "Ambiguous - same overlap":
            spn_detail = build_spn_detail(candidate_spns, bom_set, sys_groups)
        else:
            spn_detail = ""

        if not spn:
            compare_rows.append({
                "Customer_CPN": cpn,
                "SPN": "",
                "Description": desc,
                "Location": loc,
                "Candidate_SPNs": candidate_spns,
                "Selection_Status": selection_status,
                "BOM_MPN_List": " / ".join(sorted(bom_set)),
                "System_MPN_List": "",
                "Crossed_MPN_List": "",
                "Missing_In_System": "",
                "Extra_In_System": "",
                "SPN_Detail_Comparison": spn_detail,
                "MPN_Compare_Status": selection_status if selection_status else "Missing SPN",
            })
            continue

        system_set = sys_groups.get(spn, set())
        crossed_set = crossed_groups.get(spn, set())
        missing_in_system = sorted(bom_set - system_set)
        extra_in_system = sorted(system_set - bom_set)
        # BOM 用到的 MPN 中，有哪些在系統裡是被 cross 的（最優先標注）。
        crossed_in_bom = sorted(bom_set & crossed_set)

        if crossed_in_bom:
            status = "MPN Crossed in System"
        elif not bom_set and not system_set:
            status = "No MPN Data"
        elif not missing_in_system and not extra_in_system:
            status = "Full Match"
        elif missing_in_system and extra_in_system:
            status = "Partial Match"
        elif missing_in_system:
            status = "Missing in System"
        else:
            status = "Extra in System"

        compare_rows.append({
            "Customer_CPN": cpn,
            "SPN": spn,
            "Description": desc,
            "Location": loc,
            "Candidate_SPNs": candidate_spns,
            "Selection_Status": selection_status,
            "BOM_MPN_List": " / ".join(sorted(bom_set)),
            "System_MPN_List": " / ".join(sorted(system_set)),
            "Crossed_MPN_List": " / ".join(crossed_in_bom),
            "Missing_In_System": " / ".join(missing_in_system),
            "Extra_In_System": " / ".join(extra_in_system),
            "SPN_Detail_Comparison": spn_detail,
            "MPN_Compare_Status": status,
        })

    return pd.DataFrame(compare_rows)

def build_missing_spn_list(mapped_df, original_mpn_df):
    mpn_grouped = (
        original_mpn_df.groupby('Customer_CPN')['BOM_MPN']
        .agg(lambda x: ' / '.join(sorted({normalize_key(v) for v in x if normalize_key(v)})))
        .reset_index()
    )

    missing = mapped_df[
        mapped_df['Selection_Status'].isin(['Missing SPN', 'Ambiguous - same overlap'])
    ].copy()

    missing = missing.merge(mpn_grouped, on='Customer_CPN', how='left')
    missing = missing.rename(columns={'BOM_MPN': 'BOM_MPN_List'})

    suggestion_rows = []
    for _, row in missing.iterrows():
        code, category_desc, reason = suggest_part_number_category(
            row.get('Description', ''),
            row.get('Location', ''),
            row.get('BOM_MPN_List', ''),
        )

        suggestion_rows.append({
            'Customer_CPN': row.get('Customer_CPN', ''),
            'Description': row.get('Description', ''),
            'Qty_Per_Board': row.get('Qty_Per_Board', ''),
            'Location': row.get('Location', ''),
            'Candidate_SPNs': row.get('Candidate_SPNs', ''),
            'Selection_Status': row.get('Selection_Status', ''),
            'BOM_MPN_List': row.get('BOM_MPN_List', ''),
            'Suggested_Part_Number_Category': code,
            'Suggested_Category_Description': category_desc,
            'Suggested_SPN_Prefix': f'{code}-' if code else '',
            'Classification_Status': 'Suggested' if code else 'Needs Review',
            'Suggestion_Basis': reason,
        })

    return pd.DataFrame(suggestion_rows)

def build_summary(original_base_df, mapped_df, compare_df):
    return pd.DataFrame([
        {'Metric': 'Original BOM rows', 'Value': len(original_base_df)},
        {'Metric': 'Unique match count', 'Value': int((mapped_df['Selection_Status'] == 'Unique match').sum())},
        {'Metric': 'Auto-selected by MPN overlap count', 'Value': int((mapped_df['Selection_Status'] == 'Auto-selected by MPN overlap').sum())},
        {'Metric': 'Ambiguous count', 'Value': int((mapped_df['Selection_Status'] == 'Ambiguous - same overlap').sum())},
        {'Metric': 'Missing SPN count', 'Value': int((mapped_df['Selection_Status'] == 'Missing SPN').sum())},
        {'Metric': 'Full Match MPN count', 'Value': int((compare_df['MPN_Compare_Status'] == 'Full Match').sum())},
        {'Metric': 'Partial Match count', 'Value': int((compare_df['MPN_Compare_Status'] == 'Partial Match').sum())},
        {'Metric': 'Missing in System count', 'Value': int((compare_df['MPN_Compare_Status'] == 'Missing in System').sum())},
        {'Metric': 'Extra in System count', 'Value': int((compare_df['MPN_Compare_Status'] == 'Extra in System').sum())},
        {'Metric': 'MPN Crossed in System count', 'Value': int((compare_df['MPN_Compare_Status'] == 'MPN Crossed in System').sum())},
    ])

def make_result_excel(original_base_df, original_mpn_df, mapped_df, compare_df, missing_spn_df, summary_df):
    output = io.BytesIO()
    compare_sorted = compare_df.copy()

    status_priority = {
        "Missing SPN": 1,
        "Ambiguous - same overlap": 2,
        "MPN Crossed in System": 3,
        "Missing in System": 4,
        "Extra in System": 5,
        "Partial Match": 6,
        "No MPN Data": 7,
        "Full Match": 8,
    }

    compare_sorted["__sort_priority"] = compare_sorted["MPN_Compare_Status"].map(status_priority).fillna(99)
    compare_sorted = compare_sorted.sort_values(
        by=["__sort_priority", "MPN_Compare_Status", "Customer_CPN"],
        ascending=[True, True, True],
        kind="stable"
    ).drop(columns=["__sort_priority"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        original_base_df.to_excel(writer, sheet_name="Original_BOM_Normalized", index=False)
        original_mpn_df.to_excel(writer, sheet_name="Original_BOM_MPN_List", index=False)
        mapped_df.to_excel(writer, sheet_name="CPN_to_SPN_Map", index=False)
        compare_sorted.to_excel(writer, sheet_name="MPN_Compare", index=False)
        missing_spn_df.to_excel(writer, sheet_name="Missing_SPN", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        from openpyxl.styles import PatternFill

        ws = writer.book["MPN_Compare"]

        # 顏色分開
        missing_spn_fill = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")   # 最明顯深紅
        ambiguous_fill = PatternFill(fill_type="solid", start_color="E06666", end_color="E06666")     # 紅
        crossed_fill = PatternFill(fill_type="solid", start_color="B4A7D6", end_color="B4A7D6")       # 紫（被 cross）
        missing_system_fill = PatternFill(fill_type="solid", start_color="F4B183", end_color="F4B183")# 深橘
        extra_system_fill = PatternFill(fill_type="solid", start_color="FCE5CD", end_color="FCE5CD")  # 淡橘
        partial_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")       # 黃

        headers = [cell.value for cell in ws[1]]
        status_col_idx = headers.index("MPN_Compare_Status") + 1

        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=status_col_idx).value

            if status == "Missing SPN":
                fill = missing_spn_fill
            elif status == "Ambiguous - same overlap":
                fill = ambiguous_fill
            elif status == "MPN Crossed in System":
                fill = crossed_fill
            elif status == "Missing in System":
                fill = missing_system_fill
            elif status == "Extra in System":
                fill = extra_system_fill
            elif status == "Partial Match":
                fill = partial_fill
            else:
                fill = None

            if fill:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    output.seek(0)
    return output
    
