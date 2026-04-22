
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


RED_FILL = PatternFill(fill_type="solid", fgColor="FDE2E1")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF4CC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
BOLD_FONT = Font(bold=True)


@dataclass
class ProcessingResult:
    summary: Dict[str, int]
    mapping_df: pd.DataFrame
    duplicate_df: pd.DataFrame
    export_bytes: bytes


def find_header_row(raw_df: pd.DataFrame) -> int:
    """Find the row containing both Material and Customer material no."""
    for idx in range(min(30, len(raw_df))):
        row_values = [str(v).strip() for v in raw_df.iloc[idx].tolist()]
        if "Material" in row_values and "Customer material no." in row_values:
            return idx
    raise ValueError(
        "Could not find a header row containing both 'Material' and "
        "'Customer material no.'."
    )


def load_source_data(uploaded_file) -> pd.DataFrame:
    """
    Read the first worksheet from the uploaded Excel file and return only SPN/CPN data.
    """
    raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    header_row = find_header_row(raw)

    headers = raw.iloc[header_row].tolist()
    df = raw.iloc[header_row + 1 :].copy()
    df.columns = headers
    df = df.reset_index(drop=True)

    required_cols = ["Material", "Customer material no."]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    df = df[required_cols].rename(
        columns={"Material": "SPN", "Customer material no.": "CPN"}
    )

    df = df[df["SPN"].notna()].copy()
    df["SPN"] = df["SPN"].astype(str).str.strip()
    df["CPN"] = df["CPN"].fillna("").astype(str).str.strip()

    df = df[(df["SPN"] != "") & (df["CPN"] != "")]
    df = df.drop_duplicates(subset=["SPN", "CPN"]).reset_index(drop=True)
    return df


def split_cpn_field(cpn_value: str) -> List[str]:
    """Split a slash-separated CPN field into individual CPN values."""
    if pd.isna(cpn_value):
        return []

    text = str(cpn_value).strip()
    if not text:
        return []

    parts = [part.strip() for part in text.split("/") if str(part).strip()]
    return parts or [text]


def expand_spn_cpn_pairs(df: pd.DataFrame) -> pd.DataFrame:
    records = []

    for _, row in df.iterrows():
        spn = row["SPN"]
        original_cpn = row["CPN"]
        split_cpns = split_cpn_field(original_cpn)
        is_multi_cpn_spn = len(split_cpns) > 1

        for cpn in split_cpns:
            records.append(
                {
                    "SPN": spn,
                    "CPN": cpn,
                    "Original CPN Field": original_cpn,
                    "Is Multi-CPN SPN": is_multi_cpn_spn,
                }
            )

    mapping_df = pd.DataFrame(records)

    duplicate_mask = (
        mapping_df.groupby("CPN")["SPN"].transform("nunique") > 1
    )
    mapping_df["Is Duplicate CPN"] = duplicate_mask

    def build_remarks(row) -> str:
        remarks = []
        if row["Is Duplicate CPN"]:
            remarks.append("Duplicate CPN")
        if row["Is Multi-CPN SPN"]:
            remarks.append("Multi-CPN SPN")
        return "; ".join(remarks)

    mapping_df["Remarks"] = mapping_df.apply(build_remarks, axis=1)
    mapping_df = mapping_df.sort_values(
        by=["Is Duplicate CPN", "Is Multi-CPN SPN", "CPN", "SPN"],
        ascending=[False, False, True, True],
    ).reset_index(drop=True)

    return mapping_df


def build_duplicate_cpn_table(mapping_df: pd.DataFrame) -> pd.DataFrame:
    duplicate_only = mapping_df[mapping_df["Is Duplicate CPN"]].copy()
    if duplicate_only.empty:
        return pd.DataFrame(columns=["CPN", "Note"])

    grouped = (
        duplicate_only.groupby("CPN")["SPN"]
        .apply(lambda s: sorted(set(s)))
        .reset_index()
    )

    max_spn_count = grouped["SPN"].apply(len).max()
    rows = []

    for _, row in grouped.iterrows():
        spns = row["SPN"]
        output_row = {"CPN": row["CPN"]}
        for i, spn in enumerate(spns, start=1):
            output_row[f"SPN {i}"] = spn
        output_row["Note"] = f"{len(spns)} SPNs share this CPN"
        rows.append(output_row)

    duplicate_df = pd.DataFrame(rows)
    ordered_columns = ["CPN"] + [f"SPN {i}" for i in range(1, max_spn_count + 1)] + ["Note"]
    duplicate_df = duplicate_df.reindex(columns=ordered_columns)
    return duplicate_df.sort_values("CPN").reset_index(drop=True)


def build_summary(source_df: pd.DataFrame, mapping_df: pd.DataFrame) -> Dict[str, int]:
    multi_spn_count = source_df["CPN"].astype(str).str.contains("/", regex=False).sum()

    duplicate_cpn_counts = mapping_df.groupby("CPN")["SPN"].nunique()
    duplicate_cpn_values = duplicate_cpn_counts[duplicate_cpn_counts > 1]
    duplicate_rows_affected = mapping_df[mapping_df["CPN"].isin(duplicate_cpn_values.index)].shape[0]

    return {
        "Total SPN records (original)": int(len(source_df)),
        "Total expanded SPN-CPN pairs": int(len(mapping_df)),
        "Unique SPNs": int(source_df["SPN"].nunique()),
        "Unique CPNs": int(mapping_df["CPN"].nunique()),
        "SPNs with multiple CPNs": int(multi_spn_count),
        "Duplicate CPN values": int(len(duplicate_cpn_values)),
        "Rows affected by duplicate CPNs": int(duplicate_rows_affected),
    }


def auto_fit_worksheet(ws) -> None:
    for col_cells in ws.columns:
        max_length = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)


def create_export_workbook(
    summary: Dict[str, int], mapping_df: pd.DataFrame, duplicate_df: pd.DataFrame
) -> bytes:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "CPN Mapping Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)

    row_idx = 3
    for key, value in summary.items():
        ws_summary.cell(row=row_idx, column=1, value=key)
        ws_summary.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    for cell in ws_summary[3]:
        cell.font = BOLD_FONT
    auto_fit_worksheet(ws_summary)

    ws_mapping = wb.create_sheet("SPN-CPN Mapping")
    mapping_columns = ["SPN", "CPN", "Remarks"]
    ws_mapping.append(mapping_columns)
    for cell in ws_mapping[1]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL

    for _, row in mapping_df.iterrows():
        ws_mapping.append([row["SPN"], row["CPN"], row["Remarks"]])

    for row in range(2, ws_mapping.max_row + 1):
        remark_text = str(ws_mapping.cell(row=row, column=3).value or "")
        if "Duplicate CPN" in remark_text:
            for col in range(1, 4):
                ws_mapping.cell(row=row, column=col).fill = RED_FILL
        elif "Multi-CPN SPN" in remark_text:
            for col in range(1, 4):
                ws_mapping.cell(row=row, column=col).fill = YELLOW_FILL

    auto_fit_worksheet(ws_mapping)

    ws_dup = wb.create_sheet("Duplicate CPNs")
    if duplicate_df.empty:
        ws_dup.append(["CPN", "Note"])
        ws_dup.append(["", "No duplicate CPNs found"])
    else:
        ws_dup.append(list(duplicate_df.columns))
        for cell in ws_dup[1]:
            cell.font = BOLD_FONT
            cell.fill = HEADER_FILL

        for _, row in duplicate_df.iterrows():
            ws_dup.append(row.tolist())

    auto_fit_worksheet(ws_dup)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def process_uploaded_file(uploaded_file) -> ProcessingResult:
    source_df = load_source_data(uploaded_file)
    mapping_df = expand_spn_cpn_pairs(source_df)
    duplicate_df = build_duplicate_cpn_table(mapping_df)
    summary = build_summary(source_df, mapping_df)
    export_bytes = create_export_workbook(summary, mapping_df, duplicate_df)

    return ProcessingResult(
        summary=summary,
        mapping_df=mapping_df,
        duplicate_df=duplicate_df,
        export_bytes=export_bytes,
    )
